import base64
import cryptography
import pandas as pd
from openpyxl import load_workbook
from openpyxl.packaging.custom import StringProperty

import config
import hash
import password

def insert_fio_to_ws(ws, fio):
    ws['A1'] = fio

def get_fio_from_ws(ws):
    fio = ws['A1'].value
    if fio is None:
        fio = ''
    return fio

def get_fio_from_properties(path):
    file_hash = hash.from_excel_file(path)
    key = base64.urlsafe_b64encode(file_hash.digest())

    wb = load_workbook(path)
    prop = wb.custom_doc_props["fio"]
    return password.decrypt(prop.value, key).decode()


def check_file(path):
    wb = load_workbook(path)
    ws = wb.active

    fio_in_file = get_fio_from_ws(ws)
    fio_in_properties = get_fio_from_properties(path)

    if fio_in_file == fio_in_properties:
        return fio_in_file
    else:
        return ''

def excel_test_data_to_file(data, fio, path):
    test_ok = []
    bally = []

    for test_res in data:
        test_ok.append(test_res['Тест пройден'])
        bally.append(test_res['Количество баллов'])

    df = pd.DataFrame({
        'Тест пройден': test_ok,
        'Количество баллов': bally
    })
    df.index = range(1, len(df) + 1)  # Чтобы нумерация была не с 0, а с 1
    df.to_excel(path)

    wb = load_workbook(path)
    ws = wb.active
    insert_fio_to_ws(ws, fio)
    wb.save(path)

def make_and_write_secret_to_metadata(path, fio):
    file_hash = hash.from_excel_file(path)

    key = base64.urlsafe_b64encode(file_hash.digest())
    token = password.encrypt(fio.encode(), key)

    wb = load_workbook(path)
    wb.custom_doc_props.append(StringProperty(name="fio", value=token.decode()))

    wb.save(path)

if __name__ == "__main__":
    excel_data = [
        {
            'Тест пройден': 'Да',
            'Количество баллов': 5
        },
        {
            'Тест пройден': 'Да',
            'Количество баллов': 4
        },
        {
            'Тест пройден': 'Нет',
            'Количество баллов': 2
        },
        {
            'Тест пройден': 'Нет',
            'Количество баллов': 2
        },
        {
            'Тест пройден': 'Да',
            'Количество баллов': 3
        }
    ]

    fio = 'Ибрагимов Руслан Сиражутинович'

    excel_test_data_to_file(excel_data, fio, 'res.xlsx')

    make_and_write_secret_to_metadata('res.xlsx', fio)
